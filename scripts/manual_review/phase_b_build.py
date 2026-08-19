"""
Phase B: 340 페어 전수 review 준비
1. phase_b_images/ — 비교 이미지 340장 (우선순위 순 파일명)
2. sfa_review_master.xlsx — 직접 수정 가능한 마스터 파일
"""
import csv, os, sys, shutil
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
os.chdir(_REPO_ROOT)

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

import os as _os
from pathlib import Path as _Path
# 저장소 루트는 이 파일 위치에서 유도한다(개인 경로 하드코딩 금지).
# 다른 위치에서 쓰려면 HVF_ROOT 환경변수로 덮어쓴다.
_REPO_ROOT = _os.environ.get('HVF_ROOT', str(_Path(__file__).resolve().parents[2]))


# ── 상수 ────────────────────────────────────────────────
LAYOUT_OD = {0:[2,3,4,5], 1:[1,2,3,4,5,6], 2:[0,1,2,3,4,5,6,7],
             3:[0,1,2,3,4,5,6,7],   4:[0,1,2,3,4,5,6,7],
             5:[0,1,2,3,4,5,6,7], 6:[1,2,3,4,5,6], 7:[2,3,4,5]}
LAYOUT_OS = {0:[2,3,4,5], 1:[1,2,3,4,5,6], 2:[0,1,2,3,4,5,6,7],
             3:[0,1,2,3,4,5,6,7,8], 4:[0,1,2,3,4,5,6,7,8],
             5:[0,1,2,3,4,5,6,7], 6:[1,2,3,4,5,6], 7:[2,3,4,5]}
def make_gp(layout): return [(r,c) for r in range(8) for c in layout[r]]
GP = {'OD': make_gp(LAYOUT_OD), 'OS': make_gp(LAYOUT_OS)}
PT = [f'p{i+1:02d}' for i in range(54)]
BLIND = {'OD': 'p33', 'OS': 'p29'}

# ── 데이터 ──────────────────────────────────────────────
ml = list(csv.DictReader(open('ml_dataset.csv', encoding='utf-8-sig')))
ml_map = {(r['patient_id'], r['eye'], r['vf_date']): r['sfa_dir']
          for r in ml if r.get('sfa_dir')}

rows = list(csv.DictReader(open('sfa_thresholds.csv', encoding='utf-8-sig')))
print(f'전체 페어: {len(rows)}')

# ── 의심 셀 카탈로그 ────────────────────────────────────
def grid_dict(row, eye):
    g = {}
    for i, (rr, cc) in enumerate(GP[eye]):
        v = row.get(PT[i], '')
        if v in ('', None): continue
        try: g[(rr,cc)] = int(float(v))
        except: pass
    return g

def find_suspects(row):
    """행에서 의심 셀 목록 반환: [(point, value, reason), ...]"""
    eye = row['eye']
    g = grid_dict(row, eye)
    suspects = []

    # >40 dB
    for p in PT:
        v = row.get(p, '')
        if v in ('', '-1', None): continue
        try:
            iv = int(float(v))
            if iv > 40:
                suspects.append((p, iv, 'outlier_>40'))
        except: pass

    # 인접 격차 ≥18 (blind spot 제외)
    blind_pos = {'OD': (4, 6), 'OS': (4, 1)}[eye]
    for (rr, cc), val in g.items():
        if val == -1 or (rr, cc) == blind_pos:
            continue
        nbs = [g.get((rr+dr, cc+dc)) for dr,dc in [(-1,0),(1,0),(0,-1),(0,1)]]
        nbs = [v for v in nbs if v is not None and v != -1]
        if not nbs: continue
        avg = sum(nbs)/len(nbs)
        if avg - val >= 18:
            i_p = next((i for i,(a,b) in enumerate(GP[eye]) if (a,b)==(rr,cc)), None)
            p = PT[i_p] if i_p is not None else '?'
            suspects.append((p, val, f'nb_diff_{round(avg-val,0):.0f}'))

    # blind spot 인접 35-37 dB
    if eye == 'OD':
        for p in ['p32','p34','p25','p41']:
            v = row.get(p, '')
            try:
                iv = int(float(v))
                if iv in (35,36,37):
                    suspects.append((p, iv, 'boundary_35_37'))
            except: pass
    else:
        for p in ['p28','p30','p20','p38']:
            v = row.get(p, '')
            try:
                iv = int(float(v))
                if iv in (35,36,37):
                    suspects.append((p, iv, 'boundary_35_37'))
            except: pass

    return suspects

suspects_by_pair = {}
for r in rows:
    key = (r['patient_id'], r['eye'], r['vf_date'])
    susp = find_suspects(r)
    suspects_by_pair[key] = susp

# 우선순위: outlier > nb_diff (큰 순) > boundary > clean
def priority(susp):
    if not susp: return (3, 0)
    if any(s[2]=='outlier_>40' for s in susp): return (0, len(susp))
    nb = [s for s in susp if 'nb_diff' in s[2]]
    if nb:
        max_diff = max(int(s[2].replace('nb_diff_','')) for s in nb)
        return (1, -max_diff)  # 큰 diff부터
    return (2, -len(susp))

ordered = sorted(rows, key=lambda r: priority(suspects_by_pair[(r['patient_id'], r['eye'], r['vf_date'])]))
print(f'우선순위 정렬 완료')

n_outlier = sum(1 for r in rows if any(s[2]=='outlier_>40'
                for s in suspects_by_pair[(r['patient_id'], r['eye'], r['vf_date'])]))
n_nbdiff  = sum(1 for r in rows if any('nb_diff' in s[2]
                for s in suspects_by_pair[(r['patient_id'], r['eye'], r['vf_date'])]))
n_bndry   = sum(1 for r in rows if any(s[2]=='boundary_35_37'
                for s in suspects_by_pair[(r['patient_id'], r['eye'], r['vf_date'])]))
n_clean   = sum(1 for r in rows if not suspects_by_pair[(r['patient_id'], r['eye'], r['vf_date'])])
print(f'  outlier 포함: {n_outlier}')
print(f'  nb_diff 포함: {n_nbdiff}')
print(f'  boundary 포함: {n_bndry}')
print(f'  clean (의심 신호 없음): {n_clean}')

# ── 이미지 생성 ─────────────────────────────────────────
out_dir = 'phase_b_images'
os.makedirs(out_dir, exist_ok=True)

from ocr_threshold import get_grid_info

def visualize_pair(rank, row, susp, out_path):
    pid, eye, date = row['patient_id'], row['eye'], row['vf_date']
    sfa_dir = ml_map.get((pid, eye, date), '')
    img_path = os.path.join(sfa_dir, 'threshold_grid.png') if sfa_dir else ''

    susp_points = {s[0]: s[2] for s in susp}

    fig, axes = plt.subplots(1, 2, figsize=(15, 5.8))
    title = f'#{rank:03d}  {pid}  {eye}  {date}'
    if susp:
        cats = sorted(set(s[2].split('_')[0] for s in susp))
        title += f'  [의심: {len(susp)}개 / {", ".join(cats)}]'
    else:
        title += '  [clean]'
    fig.suptitle(title, fontsize=11, fontweight='bold',
                 color='red' if susp else 'black')

    # 좌: 원본
    if img_path and os.path.isfile(img_path):
        img = Image.open(img_path)
        axes[0].imshow(img)
        axes[0].set_title('원본 threshold_grid.png')
        axes[0].axis('off')

        if susp:
            iw, ih = img.size
            row_c, col_c, half, _ = get_grid_info(iw, ih)
            for p, val, reason in susp:
                i_pt = PT.index(p)
                if i_pt < len(GP[eye]):
                    rr, cc = GP[eye][i_pt]
                    cx, cy = col_c[cc], row_c[rr]
                    color = 'red' if 'outlier' in reason else ('orange' if 'nb_diff' in reason else 'gold')
                    rect = patches.Rectangle((cx-half, cy-half), 2*half, 2*half,
                                            linewidth=2, edgecolor=color,
                                            facecolor='none', linestyle='--')
                    axes[0].add_patch(rect)
                    axes[0].annotate(f'{p}={val}', xy=(cx, cy-half-3),
                                     color=color, fontsize=7, fontweight='bold',
                                     ha='center')
    else:
        axes[0].text(0.5, 0.5, '이미지 없음', ha='center', va='center')
        axes[0].axis('off')

    # 우: OCR grid
    ax = axes[1]
    ax.set_xlim(-0.5, 8.5); ax.set_ylim(-0.5, 7.5)
    ax.invert_yaxis(); ax.set_aspect('equal'); ax.axis('off')
    ax.set_title('OCR 값 grid (의심=색칠, BS=blind spot 강제 -1)')

    blind_p = BLIND[eye]
    for i, (rr, cc) in enumerate(GP[eye]):
        p = PT[i]
        v = row.get(p, '')
        try: iv = int(float(v)) if v not in ('','-1') else (-1 if v=='-1' else None)
        except: iv = None

        is_blind = (p == blind_p)
        is_susp  = (p in susp_points)
        susp_reason = susp_points.get(p, '')

        if is_blind:
            bg = '#FFD4D4'; ec='#CC4444'; lw=1.2
            txt = 'BS'
        elif is_susp:
            if 'outlier' in susp_reason: bg='#FF6B6B'; ec='red'; lw=2
            elif 'nb_diff' in susp_reason: bg='#FFB366'; ec='orange'; lw=2
            else: bg='#FFE066'; ec='gold'; lw=2
            txt = '<0' if iv==-1 else str(iv)
        else:
            bg='#F0F8FF'; ec='#aaa'; lw=0.5
            txt = '?' if iv is None else ('<0' if iv==-1 else str(iv))

        rect = patches.FancyBboxPatch((cc-0.44, rr-0.44), 0.88, 0.88,
                                      boxstyle='round,pad=0.03',
                                      facecolor=bg, edgecolor=ec, linewidth=lw)
        ax.add_patch(rect)
        ax.text(cc, rr-0.18, p, ha='center', va='center', fontsize=5, color='#444')
        ax.text(cc, rr+0.18, txt, ha='center', va='center',
                fontsize=7.5, fontweight='bold' if (is_susp or is_blind) else 'normal',
                color='white' if (is_susp and 'outlier' in susp_reason) else 'black')

    fig.tight_layout()
    fig.savefig(out_path, dpi=110, bbox_inches='tight')
    plt.close(fig)

print('\n이미지 생성 중...')
image_paths = {}
for rank, row in enumerate(ordered, 1):
    pid, eye, date = row['patient_id'], row['eye'], row['vf_date']
    susp = suspects_by_pair[(pid, eye, date)]
    fn = f'{rank:03d}_{pid}_{eye}_{date}.png'
    p = os.path.join(out_dir, fn)
    visualize_pair(rank, row, susp, p)
    image_paths[(pid, eye, date)] = fn
    if rank % 50 == 0:
        print(f'  {rank}/{len(ordered)} 완료')

print(f'  → {out_dir}/ ({len(ordered)}개)')

# ── Excel 마스터 파일 생성 ──────────────────────────────
print('\nExcel 마스터 파일 생성 중...')
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Review'

# 헤더
headers = ['rank','patient_id','eye','vf_date','image_file','n_suspect',
           'suspect_summary','status','reviewer_note'] + PT
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color='FFFFFF', size=9)
    cell.fill = PatternFill('solid', fgColor='2E5090')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 색상
fill_outlier = PatternFill('solid', fgColor='FFB3B3')  # 빨강
fill_nbdiff  = PatternFill('solid', fgColor='FFCC80')  # 주황
fill_bndry   = PatternFill('solid', fgColor='FFF2CC')  # 노랑
fill_blind   = PatternFill('solid', fgColor='E0E0E0')  # 회색 (blind)
fill_neg     = PatternFill('solid', fgColor='F0F0F0')  # 옅은 회색 (-1)
fill_clean   = PatternFill('solid', fgColor='F0F8FF')  # 하늘색
fill_empty   = PatternFill('solid', fgColor='FFFACD')  # 미검출

thin = Side(border_style='thin', color='CCCCCC')
border = Border(top=thin, bottom=thin, left=thin, right=thin)

for ri, row in enumerate(ordered, 2):
    pid, eye, date = row['patient_id'], row['eye'], row['vf_date']
    susp = suspects_by_pair[(pid, eye, date)]
    susp_points = {s[0]: s[2] for s in susp}

    n_susp = len(susp)
    summary = '; '.join(f'{s[0]}={s[1]}({s[2]})' for s in susp[:5])
    if len(susp) > 5: summary += f' +{len(susp)-5}'

    # 메타 컬럼
    base_data = [ri-1, pid, eye, date, image_paths[(pid,eye,date)],
                 n_susp, summary, 'pending', '']
    for c, v in enumerate(base_data, 1):
        cell = ws.cell(row=ri, column=c, value=v)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=9)

    # 의심도에 따라 status 셀 색칠
    status_cell = ws.cell(row=ri, column=8)
    if any('outlier' in s[2] for s in susp): status_cell.fill = fill_outlier
    elif any('nb_diff' in s[2] for s in susp): status_cell.fill = fill_nbdiff
    elif any('boundary' in s[2] for s in susp): status_cell.fill = fill_bndry
    else: status_cell.fill = fill_clean

    # p01-p54
    blind_p = BLIND[eye]
    for pi, p in enumerate(PT):
        col = 10 + pi
        v = row.get(p, '')
        try: iv = int(float(v)) if v != '' else None
        except: iv = None

        cell = ws.cell(row=ri, column=col, value=iv if iv is not None else '')
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=9)
        cell.border = border

        if p == blind_p:
            cell.fill = fill_blind
        elif p in susp_points:
            r = susp_points[p]
            if 'outlier' in r: cell.fill = fill_outlier
            elif 'nb_diff' in r: cell.fill = fill_nbdiff
            else: cell.fill = fill_bndry
        elif iv is None:
            cell.fill = fill_empty
        elif iv == -1:
            cell.fill = fill_neg

# 열 너비
widths = {1:5, 2:11, 3:4, 4:10, 5:30, 6:6, 7:55, 8:9, 9:25}
for c, w in widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w
for c in range(10, 64):
    ws.column_dimensions[get_column_letter(c)].width = 5

# 첫 행 + 첫 9개 열 고정
ws.freeze_panes = 'J2'

# 안내 시트
ws2 = wb.create_sheet('안내')
guide_lines = [
    '=== Phase B: 340 페어 전수 review ===',
    '',
    '## 사용법',
    '1. image_file 컬럼의 파일명을 phase_b_images/ 폴더에서 열기',
    '2. 좌측: 원본 threshold map / 우측: OCR 값',
    '3. p01~p54 셀에서 잘못된 값 직접 수정',
    '4. status 셀을 reviewed / skip 으로 변경',
    '5. 저장 후 → python phase_b_apply.py 실행',
    '',
    '## status 값',
    '  pending   = 미검토',
    '  reviewed  = 검토 완료 (수정 있든 없든)',
    '  skip      = 이미지 없음/판독 불가',
    '',
    '## 색상 의미 (p01~p54 셀)',
    '  빨강    = >40 dB outlier (확정 OCR 오류)',
    '  주황    = 인접 격차 ≥18 (의심)',
    '  노랑    = 35-37 dB boundary (보통 false positive)',
    '  회색    = blind spot (-1로 강제, 수정 금지)',
    '  옅은회색 = <0 dB (-1)',
    '  하늘색  = 정상 측정값',
    '  레몬    = 미검출 (None)',
    '',
    '## 진행 우선순위 (rank 컬럼 순)',
    '  1~ : outlier 포함 케이스',
    '  ~  : nb_diff 큰 순',
    '  ~  : boundary 포함',
    '  뒤 : clean 케이스 (스캔하듯 빠르게)',
    '',
    '## 팁',
    '  - 필터 활용: status=pending 으로 필터링하면 미검토만 보임',
    '  - n_suspect 컬럼으로 정렬해 의심 많은 케이스 우선',
    '  - 한 번에 50개 단위로 끊어 작업 권장',
    '  - 자주 저장 (Ctrl+S)',
]
for r, line in enumerate(guide_lines, 1):
    cell = ws2.cell(row=r, column=1, value=line)
    if line.startswith('==='): cell.font = Font(bold=True, size=12)
    elif line.startswith('## '): cell.font = Font(bold=True)
ws2.column_dimensions['A'].width = 70

out_xlsx = 'sfa_review_master.xlsx'
wb.save(out_xlsx)
print(f'  → {out_xlsx}')

# ── 진행 추적 CSV (Excel 못 열 때 backup) ────────────────
print('\n=== 완료 ===')
print(f'  이미지: {out_dir}/ ({len(ordered)}개)')
print(f'  Excel: {out_xlsx}')
print(f'  안내 시트 포함')
print()
print(f'우선순위 분포:')
print(f'  outlier 포함:  {n_outlier}건 (rank 상위)')
print(f'  nb_diff 포함:  {n_nbdiff}건')
print(f'  boundary 포함: {n_bndry}건')
print(f'  clean:         {n_clean}건 (rank 하위, 빠르게 스캔)')
